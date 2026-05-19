"""XLSX text extraction using openpyxl."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def extract(file_path: Path) -> str:
    """Extract cell data from all sheets in an XLSX file."""
    wb = load_workbook(str(file_path), read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c for c in cells):
                rows.append(" | ".join(cells))
        if rows:
            parts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(parts)
